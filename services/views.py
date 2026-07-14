"""
Views for services app
"""

from rest_framework import status, permissions
from rest_framework.response import Response
from rest_framework.views import APIView
from vehicles.models import Vehicle
from .models import ServiceHistory
from .serializers import ServiceHistorySerializer
import io
from django.http import HttpResponse


class ServiceHistoryListCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # Get service history for all user's vehicles
        vehicle_id = request.query_params.get('vehicle_id', '')
        if vehicle_id:
            try:
                vehicle = Vehicle.objects.get(id=vehicle_id, owner=request.user)
                services = ServiceHistory.objects.filter(vehicle=vehicle)
            except Vehicle.DoesNotExist:
                return Response({'detail': 'Vehicle not found.'}, status=status.HTTP_404_NOT_FOUND)
        else:
            services = ServiceHistory.objects.filter(vehicle__owner=request.user)

        serializer = ServiceHistorySerializer(services, many=True, context={'request': request})
        return Response({
            'success': True,
            'count': services.count(),
            'services': serializer.data
        })

    def post(self, request):
        vehicle_id = request.data.get('vehicle')
        try:
            vehicle = Vehicle.objects.get(id=vehicle_id, owner=request.user)
        except Vehicle.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Vehicle not found or does not belong to you.'
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = ServiceHistorySerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            service = serializer.save()
            # Update vehicle's last service date if this is more recent
            if service.service_date >= vehicle.last_service_date:
                vehicle.last_service_date = service.service_date
                vehicle.odometer_reading = service.odometer
                vehicle.save(update_fields=['last_service_date', 'odometer_reading'])

            return Response({
                'success': True,
                'message': 'Service record added successfully.',
                'service': ServiceHistorySerializer(service, context={'request': request}).data
            }, status=status.HTTP_201_CREATED)
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


class ServiceHistoryDetailView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get_service(self, pk, user):
        try:
            return ServiceHistory.objects.get(pk=pk, vehicle__owner=user)
        except ServiceHistory.DoesNotExist:
            return None

    def get(self, request, pk):
        service = self.get_service(pk, request.user)
        if not service:
            return Response({'detail': 'Service record not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ServiceHistorySerializer(service, context={'request': request})
        return Response({'success': True, 'service': serializer.data})

    def put(self, request, pk):
        service = self.get_service(pk, request.user)
        if not service:
            return Response({'detail': 'Service record not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ServiceHistorySerializer(
            service, data=request.data, partial=True, context={'request': request}
        )
        if serializer.is_valid():
            service = serializer.save()
            return Response({
                'success': True,
                'message': 'Service record updated successfully.',
                'service': ServiceHistorySerializer(service, context={'request': request}).data
            })
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        service = self.get_service(pk, request.user)
        if not service:
            return Response({'detail': 'Service record not found.'}, status=status.HTTP_404_NOT_FOUND)
        service.delete()
        return Response({'success': True, 'message': 'Service record deleted successfully.'})


class AdminServicesView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not request.user.is_admin:
            return Response({'detail': 'Access denied.'}, status=status.HTTP_403_FORBIDDEN)
        services = ServiceHistory.objects.all().select_related('vehicle', 'vehicle__owner')
        serializer = ServiceHistorySerializer(services, many=True, context={'request': request})
        return Response({
            'success': True,
            'count': services.count(),
            'services': serializer.data
        })


class ExportServicesView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        export_format = request.query_params.get('format', 'excel')
        vehicle_id = request.query_params.get('vehicle_id', '')

        if vehicle_id:
            services = ServiceHistory.objects.filter(
                vehicle__id=vehicle_id, vehicle__owner=request.user
            ).select_related('vehicle')
        else:
            services = ServiceHistory.objects.filter(
                vehicle__owner=request.user
            ).select_related('vehicle')

        if export_format == 'excel':
            return self._export_excel(services)
        return self._export_excel(services)

    def _export_excel(self, services):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Service History'

        headers = ['Vehicle', 'Registration', 'Service Date', 'Odometer (km)', 'Cost', 'Service Center', 'Description']
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row, service in enumerate(services, 2):
            ws.cell(row=row, column=1, value=service.vehicle.name)
            ws.cell(row=row, column=2, value=service.vehicle.registration_number)
            ws.cell(row=row, column=3, value=str(service.service_date))
            ws.cell(row=row, column=4, value=service.odometer)
            ws.cell(row=row, column=5, value=float(service.cost))
            ws.cell(row=row, column=6, value=service.service_center or '')
            ws.cell(row=row, column=7, value=service.description or '')

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="service_history.xlsx"'
        return response
